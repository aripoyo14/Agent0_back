# app/services/file_analyzer.py
"""
ファイル内容解析サービス
- PDF、Word、Excel、テキストファイルの内容を抽出
- AI返信生成時にファイル内容を考慮
"""

import requests
import io
from typing import Optional, Dict, Any
from fastapi import HTTPException
import PyPDF2
from docx import Document
from openpyxl import load_workbook
import re

def extract_file_content(file_url: str, file_type: str, file_name: str) -> Dict[str, Any]:
    """
    ファイルURLから内容を抽出する
    軽量版：最初の数KBのみ読み取り、メタデータを抽出
    """
    try:
        # ファイルをダウンロード
        response = requests.get(file_url, timeout=10)
        response.raise_for_status()
        file_content = response.content
        
        # ファイルサイズ制限（最初の50KBのみ）
        MAX_SIZE = 50 * 1024
        if len(file_content) > MAX_SIZE:
            file_content = file_content[:MAX_SIZE]
        
        # ファイルタイプ別の解析
        if file_type == "application/pdf":
            return extract_pdf_content(file_content, file_name)
        elif file_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document", "application/msword"]:
            return extract_word_content(file_content, file_name)
        elif file_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
            return extract_excel_content(file_content, file_name)
        elif file_type.startswith("text/"):
            return extract_text_content(file_content, file_name)
        else:
            return {
                "file_name": file_name,
                "file_type": file_type,
                "content": "このファイル形式は現在サポートされていません。",
                "summary": "未対応ファイル形式"
            }
            
    except Exception as e:
        return {
            "file_name": file_name,
            "file_type": file_type,
            "content": f"ファイル読み取りエラー: {str(e)}",
            "summary": "読み取り失敗"
        }

def extract_pdf_content(file_content: bytes, file_name: str) -> Dict[str, Any]:
    """PDFファイルの内容を抽出"""
    try:
        pdf_file = io.BytesIO(file_content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        
        # 最初の3ページのみ抽出
        text_content = ""
        for i, page in enumerate(pdf_reader.pages[:3]):
            text_content += f"\n--- ページ {i+1} ---\n"
            text_content += page.extract_text()
        
        # 概要を生成（最初の200文字）
        summary = text_content[:200] + "..." if len(text_content) > 200 else text_content
        
        return {
            "file_name": file_name,
            "file_type": "application/pdf",
            "content": text_content,
            "summary": summary,
            "page_count": len(pdf_reader.pages)
        }
    except Exception as e:
        return {
            "file_name": file_name,
            "file_type": "application/pdf",
            "content": f"PDF読み取りエラー: {str(e)}",
            "summary": "PDF読み取り失敗"
        }

def extract_word_content(file_content: bytes, file_name: str) -> Dict[str, Any]:
    """Wordファイルの内容を抽出"""
    try:
        doc_file = io.BytesIO(file_content)
        doc = Document(doc_file)
        
        # 段落とテーブルからテキストを抽出
        text_content = ""
        for para in doc.paragraphs[:20]:  # 最初の20段落のみ
            if para.text.strip():
                text_content += para.text + "\n"
        
        # 概要を生成
        summary = text_content[:200] + "..." if len(text_content) > 200 else text_content
        
        return {
            "file_name": file_name,
            "file_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "content": text_content,
            "summary": summary
        }
    except Exception as e:
        return {
            "file_name": file_name,
            "file_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "content": f"Word読み取りエラー: {str(e)}",
            "summary": "Word読み取り失敗"
        }

def extract_excel_content(file_content: bytes, file_name: str) -> Dict[str, Any]:
    """Excelファイルの内容を抽出"""
    try:
        excel_file = io.BytesIO(file_content)
        wb = load_workbook(excel_file, data_only=True)
        
        # 最初のシートのみ抽出
        ws = wb.active
        text_content = f"シート名: {ws.title}\n\n"
        
        # 最初の10行×10列のデータを抽出
        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
            row_data = []
            for cell in row:
                if cell.value:
                    row_data.append(str(cell.value))
            if row_data:
                text_content += " | ".join(row_data) + "\n"
        
        summary = f"Excelファイル: {ws.title}シートのデータ"
        
        return {
            "file_name": file_name,
            "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "content": text_content,
            "summary": summary
        }
    except Exception as e:
        return {
            "file_name": file_name,
            "file_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "content": f"Excel読み取りエラー: {str(e)}",
            "summary": "Excel読み取り失敗"
        }

def extract_text_content(file_content: bytes, file_name: str) -> Dict[str, Any]:
    """テキストファイルの内容を抽出"""
    try:
        text_content = file_content.decode('utf-8', errors='ignore')
        
        # 最初の1000文字のみ
        if len(text_content) > 1000:
            text_content = text_content[:1000] + "..."
        
        summary = text_content[:200] + "..." if len(text_content) > 200 else text_content
        
        return {
            "file_name": file_name,
            "file_type": "text/plain",
            "content": text_content,
            "summary": summary
        }
    except Exception as e:
        return {
            "file_name": file_name,
            "file_type": "text/plain",
            "content": f"テキスト読み取りエラー: {str(e)}",
            "summary": "テキスト読み取り失敗"
        } 